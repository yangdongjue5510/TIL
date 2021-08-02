#load_workbook is always load all file of xlsx file. this method can make problem.

from openpyxl import load_workbook

wb = load_workbook('test_data.xlsx', read_only=True)#we should append read only param. we load some cells from file.
data = wb.active

for row in data.iter_rows():
    for cell in row:
        print(cell.value)