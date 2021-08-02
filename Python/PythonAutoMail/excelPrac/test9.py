from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet('sheet_test2')

ws['A1']= 'yang'
ws['B2']= 'dongjue'

wb.save('result.xlsx')